import openpyxl
import re
from openpyxl.styles import PatternFill


def excel_cleaner(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    sheets = wb.sheetnames
    report_ws = wb.create_sheet('FIX REPORT')

    cyrillic_re = re.compile(r'[А-Яа-яЁё]')
    latin_re = re.compile(r'[A-Za-z]')
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    counter_strip = 0
    counter_cyr = 0
    for sheet in sheets:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=0, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for n, cell in enumerate(row):
                old_value = cell.value
                if n < 2 and isinstance(cell.value, str) and cell.value.startswith('Таблица'):
                    t, t_num = cell.value.strip().split(' ')
                    t_num1, t_num2 = t_num.split('.')[:-1], t_num.split('.')[-1]
                    if len(t_num2) == 1:
                        t_num = f'{'.'.join(t_num1)}.{t_num2.zfill(2)}'
                        cell.value = t + ' ' + t_num
                        report_ws.append([f'[Изменено наименование таблицы] #### "{old_value}" ----> "{cell.value}"'])

                elif cell.value and isinstance(cell.value, str):
                    if cell.value.strip() != old_value:
                        cell.value = cell.value.strip()
                        report_ws.append([
                            f'[Удалены незначащие символы в начале и конце строки] #### "{old_value}" ----> "{cell.value}"'])
                        counter_strip += 1

                value = str(cell.value) if cell.value is not None else ""
                words = re.findall(r'\b\w+\b', value)
                for word in words:
                    if cyrillic_re.search(word) and latin_re.search(word):
                        cell.fill = red_fill
                        cyr_letter = cyrillic_re.search(word).group()
                        lat_letter = latin_re.search(word).group()
                        report_ws.append([f'[Найдено некорректное слово] #### "{word}" (Кириллица: "{cyr_letter}", Латиница: "{lat_letter}"), ОТМЕЧЕНО КРАСНЫМ'])
                        counter_cyr += 1

    wb.save(out_path)
    return f'Удалены лишние символы в {counter_strip} ячейках.\nНайдено {counter_cyr} ячеек с CYR+LATIN словами.'


if __name__ == '__main__':
    excel_cleaner(r'C:\Python\Projects\AtomskillsCleaner\ADD_D.xlsx', '123.xlsx')
